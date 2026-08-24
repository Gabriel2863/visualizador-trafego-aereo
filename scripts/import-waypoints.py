#!/usr/bin/env python3
"""Converte waypoint.xlsx para uma base JSON consumida pelo mapa.

Uso: python scripts/import-waypoints.py [arquivo.xlsx] [diretorio-de-saida]
"""
import json
import math
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "waypoint.xlsx"
OUTPUT_DIR = Path(sys.argv[2]) if len(sys.argv) > 2 else ROOT / "data"


def json_value(value):
    """Converte somente tipos não serializáveis; não normaliza valores da planilha."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def main():
    workbook = load_workbook(INPUT, read_only=True, data_only=False)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows())]

    required = {"ident", "latitude", "longitude", "tipo"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"Colunas obrigatórias ausentes: {', '.join(sorted(missing))}")

    features = []
    invalid_records = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        properties = {header: json_value(value) for header, value in zip(headers, row)}
        ident = properties.get("ident")
        latitude = properties.get("latitude")
        longitude = properties.get("longitude")
        reasons = []

        if ident is None or str(ident).strip() == "":
            reasons.append("ident ausente")
        if not isinstance(latitude, (int, float)) or isinstance(latitude, bool) or not math.isfinite(latitude):
            reasons.append("latitude ausente ou não numérica")
        elif not -90 <= latitude <= 90:
            reasons.append("latitude fora do intervalo [-90, 90]")
        if not isinstance(longitude, (int, float)) or isinstance(longitude, bool) or not math.isfinite(longitude):
            reasons.append("longitude ausente ou não numérica")
        elif not -180 <= longitude <= 180:
            reasons.append("longitude fora do intervalo [-180, 180]")

        if reasons:
            invalid_records.append({"sourceRow": row_number, "reasons": reasons, "properties": properties})
            continue

        features.append({
            "type": "Feature",
            "id": f"waypoint:{properties.get('pk', row_number)}",
            "geometry": {"type": "Point", "coordinates": [longitude, latitude]},
            "properties": properties,
            "sourceRow": row_number,
        })

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    dataset = {
        "schemaVersion": 1,
        "source": {"file": INPUT.name, "sheet": sheet.title, "headers": headers},
        "recordCount": len(features),
        "features": features,
    }
    report = {
        "source": {"file": INPUT.name, "sheet": sheet.title},
        "totalRows": sheet.max_row - 1,
        "imported": len(features),
        "ignored": len(invalid_records),
        "invalidRecords": invalid_records,
    }
    (OUTPUT_DIR / "waypoints.json").write_text(json.dumps(dataset, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    (OUTPUT_DIR / "waypoints-import-report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Importados: {len(features)}")
    print(f"Ignorados: {len(invalid_records)}")


if __name__ == "__main__":
    main()
